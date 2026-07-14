"""
Generador de reporte Excel con 3 hojas: Ventas, Inventario, Clientes.
"""
import io
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


AZUL_HEADER = "003366"
GRIS_FILA_PAR = "F2F2F2"
VERDE = "006400"
ROJO = "CC0000"


def _header_style(ws, row, cols):
    fill = PatternFill(fill_type="solid", fgColor=AZUL_HEADER)
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    alin = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alin


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def generar_reporte_excel(fecha_inicio=None, fecha_fin=None) -> bytes:
    """
    Genera un Excel con 3 hojas y retorna bytes.
    """
    from apps.pedidos.models import Pedido, ItemPedido
    from apps.inventario.models import MateriaPrima, MovimientoInventario
    from apps.clientes.models import Cliente

    wb = Workbook()

    # ── Hoja 1: Ventas ──────────────────────────────────────────────────
    ws_ventas = wb.active
    assert ws_ventas is not None  # Workbook() siempre crea una hoja activa
    ws_ventas.title = "Ventas"
    headers_ventas = [
        "Nro Pedido", "Cliente", "Estado", "Fecha Creación",
        "Fecha Entrega", "Subtotal", "IVA", "Total",
    ]
    ws_ventas.append(headers_ventas)
    _header_style(ws_ventas, 1, len(headers_ventas))

    pedidos_qs = Pedido.objects.select_related("cliente").order_by("-fecha_creacion")
    if fecha_inicio:
        pedidos_qs = pedidos_qs.filter(fecha_creacion__date__gte=fecha_inicio)
    if fecha_fin:
        pedidos_qs = pedidos_qs.filter(fecha_creacion__date__lte=fecha_fin)

    total_ventas = Decimal("0.00")
    for i, p in enumerate(pedidos_qs, 2):
        ws_ventas.append([
            p.numero,
            p.cliente.nombre_completo,
            p.get_estado_display(),
            p.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            p.fecha_promesa_entrega.strftime("%d/%m/%Y") if p.fecha_promesa_entrega else "—",
            float(p.subtotal),
            float(p.iva),
            float(p.total),
        ])
        if i % 2 == 0:
            fill = PatternFill(fill_type="solid", fgColor=GRIS_FILA_PAR)
            for col in range(1, 9):
                ws_ventas.cell(row=i, column=col).fill = fill
        total_ventas += p.total

    fila_total = ws_ventas.max_row + 1
    ws_ventas.cell(row=fila_total, column=7, value="TOTAL:")
    ws_ventas.cell(row=fila_total, column=7).font = Font(bold=True)
    ws_ventas.cell(row=fila_total, column=8, value=float(total_ventas))
    ws_ventas.cell(row=fila_total, column=8).font = Font(bold=True, color=VERDE)
    _auto_width(ws_ventas)

    # ── Hoja 2: Inventario ───────────────────────────────────────────────
    ws_inv = wb.create_sheet("Inventario")
    headers_inv = [
        "Materia Prima", "Unidad", "Stock Actual",
        "Stock Mínimo", "¿Alerta?", "Proveedor",
    ]
    ws_inv.append(headers_inv)
    _header_style(ws_inv, 1, len(headers_inv))

    for i, mp in enumerate(MateriaPrima.objects.select_related("proveedor").order_by("nombre"), 2):
        alerta = mp.stock_actual <= mp.stock_minimo
        ws_inv.append([
            mp.nombre,
            mp.unidad,
            float(mp.stock_actual),
            float(mp.stock_minimo),
            "SÍ" if alerta else "No",
            mp.proveedor.nombre if mp.proveedor else "—",
        ])
        if alerta:
            ws_inv.cell(row=i, column=5).font = Font(color=ROJO, bold=True)
        if i % 2 == 0:
            fill = PatternFill(fill_type="solid", fgColor=GRIS_FILA_PAR)
            for col in range(1, 7):
                ws_inv.cell(row=i, column=col).fill = fill
    _auto_width(ws_inv)

    # ── Hoja 3: Clientes ────────────────────────────────────────────────
    ws_cli = wb.create_sheet("Clientes")
    headers_cli = [
        "Nombre", "Cédula/RUC", "Tipo", "Teléfono",
        "Email", "Total Pedidos", "Total Compras",
    ]
    ws_cli.append(headers_cli)
    _header_style(ws_cli, 1, len(headers_cli))

    clientes_qs = Cliente.objects.filter(activo=True).order_by("nombre_completo")
    for i, cli in enumerate(clientes_qs, 2):
        total_pedidos = cli.pedidos.count()
        total_compras = sum(p.total for p in cli.pedidos.all()) or Decimal("0.00")
        ws_cli.append([
            cli.nombre_completo,
            cli.cedula_ruc,
            cli.get_tipo_display(),
            cli.telefono,
            cli.email or "—",
            total_pedidos,
            float(total_compras),
        ])
        if i % 2 == 0:
            fill = PatternFill(fill_type="solid", fgColor=GRIS_FILA_PAR)
            for col in range(1, 8):
                ws_cli.cell(row=i, column=col).fill = fill
    _auto_width(ws_cli)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
